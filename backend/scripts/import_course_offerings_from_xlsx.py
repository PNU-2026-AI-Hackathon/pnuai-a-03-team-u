"""학교 공식 배포 "학부 개설강좌 일람표" xlsx → course_offerings + course_times 적재.

크롤러 CSV(scripts/import_course_offerings.py)와 나란히 존재하는 별도 경로. 학교가 학기마다
공식 배포하는 xlsx는 다음 두 가지 장점이 있어 별도 스크립트로 처리한다:

1. **교양영역명 컬럼** 이 있어 세부 카테고리(효원핵심교양/효원균형교양/효원창의교양/기초교양/
   일반선택 등)로 courses.category를 채울 수 있다. 이 세부 값은 로드맵 채팅 에이전트의
   category 필터가 이미 인지하는 어휘라 별도 스키마 변경 없이 즉시 활용된다.
2. 학교 공식 소스라 크롤러 결과와 차이가 있을 때 진위 판정 기준.

**xlsx 컬럼 → DB 매핑** (헤더는 행 6 기준):
- 교과목번호 → Course.course_code / CourseOffering 매칭 키
- 교과목명 → Course.course_name
- 학점 → Course.credits
- 교과목구분 → Course.category (교양영역명이 비어 있을 때만)
- 교양영역명 → Course.category (우선순위 높음)
- 주관학과명 → Course.department_id (departments.name 정확 매칭)
- 대학명 → CourseOffering.school
- 분반 → CourseOffering.section
- 교수명 → CourseOffering.professor
- 수강제한인원 → CourseOffering.capacity
- 시간표 → CourseTime (기존 parse_timetable_raw 재활용)

**auto-create**: xlsx의 교과목번호가 아직 courses에 없으면 xlsx 정보로 Course 행을 자동 생성.
수강편람이 커리큘럼 카탈로그보다 넓기 때문에 필요(교양 등 학과 밖 개설 포함).

**멱등**: (course_id, year, semester, section) 유일 키로 offering upsert. CourseTime은 offering
당 실행별 1회 전량 삭제 후 재삽입.

실행:
    python -m scripts.import_course_offerings_from_xlsx \
        --xlsx "/path/to/개설강좌일람표.xlsx" --year 2026 --semester 2
    # dry-run (기본). --commit 붙이면 실제 반영.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.db import SessionLocal
from app.domains.academics.models import Department
from app.domains.courses.models import Course, CourseOffering, CourseTime
from app.ingestion.parsers.onestop_course_catalog import parse_timetable_raw


# xlsx의 헤더가 6번째 행에 있어 index 5부터 데이터가 시작.
_HEADER_ROW_INDEX = 5


def _semester_display(raw: str) -> str:
    """CLI로 넘어온 학기 값(1/2/summer/winter)을 DB 표기에 맞춘다."""
    mapping = {"1": "1학기", "2": "2학기", "summer": "여름계절수업", "winter": "겨울계절수업"}
    return mapping.get(str(raw).strip(), str(raw))


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _truncate(value: str | None, limit: int) -> str | None:
    """DB 컬럼 길이 제한을 넘는 값(팀티칭 시 교수명이 여러 명 이어져 100자 초과 등)을 자른다.
    잘린 사실을 알 수 있도록 뒷부분에 말줄임표를 붙인다.
    """
    if value is None or len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def _iter_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = None
    for i, row in enumerate(rows):
        if i < _HEADER_ROW_INDEX:
            continue
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not any(c for c in row):
            continue
        yield dict(zip(header, row, strict=False))


def import_xlsx(xlsx_path: Path, year: str, semester: str, db: Session, *, commit: bool) -> dict:
    stats = Counter()
    auto_created_samples: list[str] = []

    code_to_id: dict[str, int] = dict(
        db.execute(select(Course.course_code, Course.id).where(Course.course_code.is_not(None))).all()
    )
    dept_name_to_id: dict[str, int] = dict(db.execute(select(Department.name, Department.id)).all())

    times_cleared_for: set[int] = set()

    for row in _iter_rows(xlsx_path):
        stats["rows_total"] += 1

        course_code = _clean(row.get("교과목번호"))
        if not course_code:
            stats["skipped_no_course_code"] += 1
            continue

        course_name = _clean(row.get("교과목명"))
        subject_category = _clean(row.get("교과목구분"))
        detail_category = _clean(row.get("교양영역명"))
        category = detail_category or subject_category
        credits = _to_float(row.get("학점"))
        dept_name = _clean(row.get("주관학과명"))
        college = _clean(row.get("대학명"))
        grade = _clean(row.get("학년"))
        section = _truncate(_clean(row.get("분반")), 20)
        # 팀티칭 시 교수명이 여러 명 이어져 100자 초과할 수 있어 자름.
        professor = _truncate(_clean(row.get("교수명")), 100)
        capacity = _to_int(row.get("수강제한인원"))
        timetable_raw = _clean(row.get("시간표"))

        course_id = code_to_id.get(course_code)
        if course_id is None:
            if not course_name:
                stats["skipped_no_course_name"] += 1
                continue
            new_course = Course(
                course_code=_truncate(course_code, 50),
                course_name=_truncate(course_name, 255),
                category=_truncate(category, 50),
                credits=credits,
                department_id=dept_name_to_id.get(dept_name) if dept_name else None,
                year=_truncate(grade, 10),
                semester=_truncate(str(semester), 20),
            )
            db.add(new_course)
            db.flush()
            course_id = new_course.id
            code_to_id[course_code] = course_id
            stats["courses_auto_created"] += 1
            if len(auto_created_samples) < 20:
                auto_created_samples.append(f"{course_code}:{course_name} [{category}]")
        else:
            # 기존 Course의 category가 상위 표기("교양선택")뿐이면 xlsx의 세부 영역으로 보강한다.
            course = db.get(Course, course_id)
            if detail_category and course.category != detail_category:
                course.category = detail_category
                stats["courses_category_refined"] += 1

        semester_display = _semester_display(semester)
        offering = db.scalars(
            select(CourseOffering).where(
                CourseOffering.course_id == course_id,
                CourseOffering.year == str(year),
                CourseOffering.semester == semester_display,
                CourseOffering.section == section,
            )
        ).first()

        if offering is None:
            offering = CourseOffering(
                course_id=course_id,
                year=str(year),
                semester=semester_display,
                section=section,
                school=college,
            )
            db.add(offering)
            stats["offerings_created"] += 1
        else:
            stats["offerings_updated"] += 1

        offering.school = college or offering.school
        offering.professor = professor
        offering.capacity = capacity

        db.flush()

        if offering.id not in times_cleared_for:
            db.query(CourseTime).filter(CourseTime.offering_id == offering.id).delete(
                synchronize_session=False
            )
            times_cleared_for.add(offering.id)

        parsed_sessions = parse_timetable_raw(timetable_raw)
        stats["time_sessions_parsed"] += len(parsed_sessions)
        for session in parsed_sessions:
            db.add(
                CourseTime(
                    offering_id=offering.id,
                    day_of_week=session.day_of_week,
                    start_time=session.start_time,
                    end_time=session.end_time,
                    classroom=session.classroom or None,
                )
            )

    if commit:
        db.commit()
    else:
        db.rollback()

    return {
        "stats": dict(stats),
        "auto_created_course_samples": auto_created_samples,
        "committed": commit,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--year", required=True, help="예: 2026")
    parser.add_argument("--semester", required=True, help="1/2/summer/winter")
    parser.add_argument("--commit", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.xlsx.exists():
        print(f"xlsx not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    db = SessionLocal()
    try:
        result = import_xlsx(args.xlsx, args.year, args.semester, db, commit=args.commit)
    finally:
        db.close()
    import json
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
